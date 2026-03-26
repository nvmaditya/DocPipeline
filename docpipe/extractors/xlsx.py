from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook

from .base import BaseExtractor


class XlsxExtractor(BaseExtractor):
    def extract(self, file_path: str) -> List[Dict[str, Any]]:
        suffix = Path(file_path).suffix.lower()
        if suffix == ".csv":
            return self._extract_csv(file_path)
        return self._extract_xlsx(file_path)

    def _extract_csv(self, file_path: str) -> List[Dict[str, Any]]:
        lines: List[str] = []
        with open(file_path, "r", encoding="utf-8", errors="ignore", newline="") as handle:
            reader = csv.reader(handle)
            for row in reader:
                lines.append(" | ".join(col.strip() for col in row))
        text = "\n".join(lines)
        return [{"text": text, "page_number": None, "heading_context": "Sheet1"}]

    def _extract_xlsx(self, file_path: str) -> List[Dict[str, Any]]:
        wb = load_workbook(filename=file_path, data_only=True, read_only=True)
        records: List[Dict[str, Any]] = []
        for sheet in wb.worksheets:
            lines: List[str] = []
            for row in sheet.iter_rows(values_only=True):
                values = ["" if value is None else str(value).strip() for value in row]
                if any(values):
                    lines.append(" | ".join(values))
            if lines:
                records.append(
                    {
                        "text": "\n".join(lines),
                        "page_number": None,
                        "heading_context": sheet.title,
                    }
                )
        return records
